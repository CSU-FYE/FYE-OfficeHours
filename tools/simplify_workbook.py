#!/usr/bin/env python3
"""
The shape of data/office-hours.xlsx, and the one-time reshape into it.

Run once to migrate the workbook; imported afterwards by add_tutoring.py and
update_faculty_hours.py for the column specs, the instructions tab and
`sort_shifts`, so the schema is written down in exactly one place.

What changed and why:

  shifts    `program` is gone. It said "tutoring" on 73 rows and nothing on the
            other 68, and from that the site inferred the room and the course
            list. A blank cell that silently means "ENGR office hours in AV C144"
            is a bad thing to hand a human, so every row now names its own room
            outright. `mode` went with it - a location that starts with http is
            online, and there is no longer a way to mark a row online and leave
            it sitting in a classroom. `courses`, `active` and `notes` were
            empty on all 141 rows.

  people    `display_name`, `email` and `notes` were empty on all 49 rows.
            `courses` was too, so `courses_tutoring` takes the name: one column,
            the one that actually holds anything.

  settings  the four `default_location_*` keys collapse into one
            `default_location`, since a blank location is now the exception
            rather than the rule. `tutoring_room` replaces `default_program`
            and `program_order` as the one place that says which room runs on
            the people sheet's course lists.

Rows are sorted into the order the week happens in, and both sheets get a filter
row, so anyone can pull up their own name or their own day in two clicks.

Idempotent: run it twice and the second run is a no-op.

    python3 tools/simplify_workbook.py
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
BOOK = ROOT / "data" / "office-hours.xlsx"

OFFICE_ROOM = "AV C144"
TUTORING_ROOM = "AV C141"
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True)

# The two sheets that get edited are blue, the two that rarely do are grey, and the
# instructions are the odd one out — findable at a glance along the bottom of Excel.
TAB_COLORS = {"people": "1F3864", "shifts": "1F3864", "exceptions": "8497B0",
              "settings": "8497B0", "how to use this": "BF8F00"}

PEOPLE = [("name", 26), ("role", 12), ("courses", 52), ("languages", 18)]
SHIFTS = [("name", 26), ("day", 12), ("start", 11), ("end", 11), ("location", 34)]
EXCEPTIONS = [("name", 26), ("date", 14), ("start", 11), ("end", 11),
              ("type", 12), ("location", 24), ("note", 46)]

# Dropped keys, and the pair that replaces the room and programme ones.
DROP_SETTINGS = ["default_program", "program_order", "default_location_faculty",
                 "default_location_gtf", "default_location_la", "default_location_tutoring"]
ADD_SETTINGS = [("default_location", OFFICE_ROOM), ("tutoring_room", TUTORING_ROOM)]


def s(v):
    return str(v).strip() if v is not None else ""


def columns_of(ws):
    return {s(c.value).lower(): i for i, c in enumerate(ws[1], start=1) if s(c.value)}


def snapshot(cell):
    """A cell as plain data, so the sheet underneath it can be torn down and
    rebuilt without the yellow migration flags or their comments going with it."""
    rgb = s(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor else ""
    return {
        "value": cell.value,
        "fill": rgb[-6:] if rgb not in ("", "00000000") else None,
        "comment": (cell.comment.text, cell.comment.author or "migration") if cell.comment else None,
    }


def read_rows(ws):
    """Every row below the header as {column name: snapshot}."""
    cols = columns_of(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {name: snapshot(ws.cell(r, col)) for name, col in cols.items()}
        if any(s(c["value"]) for c in row.values()):
            rows.append(row)
    return rows


def rebuild_sheet(wb, title, spec, records):
    """Replace the sheet with a fresh one at the same position, laid out under
    `spec` - a list of (header, width). Each record maps a header to a snapshot
    or to a plain value."""
    index = wb.sheetnames.index(title)
    del wb[title]
    ws = wb.create_sheet(title, index)

    ws.sheet_properties.tabColor = TAB_COLORS.get(title)
    for i, (header, width) in enumerate(spec, start=1):
        cell = ws.cell(1, i, header)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    for r, record in enumerate(records, start=2):
        for i, (header, _) in enumerate(spec, start=1):
            source = record.get(header)
            if source is None:
                continue
            cell = ws.cell(r, i)
            if isinstance(source, dict):
                cell.value = source["value"]
                if source["fill"]:
                    cell.fill = PatternFill("solid", fgColor=source["fill"])
                if source["comment"]:
                    cell.comment = openpyxl.comments.Comment(*source["comment"])
            else:
                cell.value = source
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(spec))}{max(len(records) + 1, 2)}"
    return ws


def dropdown(ws, cells, formula, message, strict=True):
    dv = DataValidation(type="list", formula1=formula, allowBlank=True)
    dv.showErrorMessage = strict
    dv.error, dv.errorTitle = message, "Check this cell"
    ws.add_data_validation(dv)
    dv.add(cells)


def main():
    if not BOOK.exists():
        sys.exit(f"Not found: {BOOK}")
    wb = openpyxl.load_workbook(BOOK)

    # -- people ---------------------------------------------------------
    ws = wb["people"]
    # Second run: `courses_tutoring` has already become `courses`.
    course_key = "courses_tutoring" if "courses_tutoring" in columns_of(ws) else "courses"
    people = [
        {"name": row["name"], "role": row.get("role"),
         "courses": row.get(course_key), "languages": row.get("languages")}
        for row in read_rows(ws) if s(row["name"]["value"])
    ]
    ws = rebuild_sheet(wb, "people", PEOPLE, people)
    dropdown(ws, f"B2:B{len(people) + 200}", '"faculty,gtf,la"',
             "Role must be faculty, gtf, or la.")
    # Some course lists run to a dozen entries. Wrapping is set on the column, not
    # per cell, so it also catches whatever add_tutoring.py writes there next run.
    ws.column_dimensions["C"].alignment = Alignment(wrap_text=True, vertical="top")

    # -- shifts ---------------------------------------------------------
    shifts = []
    for row in read_rows(wb["shifts"]):
        name = s(row["name"]["value"])
        if not name:
            continue
        location = s(row["location"]["value"]) if "location" in row else ""
        if not location:
            # The room the old `program` column implied. Tutoring is the only
            # thing that ever ran anywhere but AV C144.
            program = s(row["program"]["value"]).lower() if "program" in row else ""
            location = TUTORING_ROOM if program == "tutoring" else OFFICE_ROOM
        shifts.append({
            "name": name,
            "day": s(row["day"]["value"]),
            "start": s(row["start"]["value"]),
            "end": s(row["end"]["value"]),
            "location": location,
        })

    def sort_key(shift):
        day = DAYS.index(shift["day"]) if shift["day"] in DAYS else len(DAYS)
        return (day, minutes(shift["start"]), shift["name"].lower())

    shifts.sort(key=sort_key)
    sh = rebuild_sheet(wb, "shifts", SHIFTS, shifts)
    last = len(shifts) + 200
    dropdown(sh, f"A2:A{last}", "PeopleNames",
             "Pick a name from the people sheet. Add them there first if they are new.")
    dropdown(sh, f"B2:B{last}", f'"{",".join(DAYS)}"', "Monday through Sunday.")
    # Not strict: the room list is a shortcut, not a rulebook. Anywhere else on
    # campus, or a meeting link for an online hour, has to be typeable.
    dropdown(sh, f"E2:E{last}", f'"{OFFICE_ROOM},{TUTORING_ROOM},Scott Engineering"',
             "", strict=False)
    for col in ("C", "D"):
        sh.column_dimensions[col].number_format = "@"

    # -- exceptions -----------------------------------------------------
    records = [{k: row.get(k) for k, _ in EXCEPTIONS} for row in read_rows(wb["exceptions"])]
    ex = rebuild_sheet(wb, "exceptions", EXCEPTIONS, records)
    last = len(records) + 200
    dropdown(ex, f"A2:A{last}", "PeopleNames", "Pick a name from the people sheet.")
    dropdown(ex, f"E2:E{last}", '"cancelled,added"', "Use cancelled or added.")
    for col in ("B", "C", "D"):
        ex.column_dimensions[col].number_format = "@"

    # -- settings -------------------------------------------------------
    st = wb["settings"]
    kept = [(s(st.cell(r, 1).value), st.cell(r, 2).value)
            for r in range(2, st.max_row + 1) if s(st.cell(r, 1).value)]
    kept = [(k, v) for k, v in kept if k not in DROP_SETTINGS]
    have = {k for k, _ in kept}
    for key, value in ADD_SETTINGS:
        if key not in have:
            kept.append((key, value))
    st.delete_rows(2, st.max_row)
    for r, (k, v) in enumerate(kept, start=2):
        st.cell(r, 1, k)
        st.cell(r, 2, v)
    st.freeze_panes = "A2"
    st.sheet_properties.tabColor = TAB_COLORS["settings"]

    # PeopleNames drives every name dropdown; keep it over the rows that exist.
    wb.defined_names["PeopleNames"] = openpyxl.workbook.defined_name.DefinedName(
        "PeopleNames", attr_text=f"people!$A$2:$A${len(people) + 200}")

    write_instructions(wb)
    wb.save(BOOK)

    print(f"people:  {len(people)} rows, {len(PEOPLE)} columns")
    print(f"shifts:  {len(shifts)} rows, {len(SHIFTS)} columns "
          f"({sum(1 for s_ in shifts if s_['location'] == TUTORING_ROOM)} in {TUTORING_ROOM})")
    print(f"settings: {len(kept)} keys")
    print(f"Wrote {BOOK.relative_to(ROOT)}")


def sort_shifts(ws):
    """Put the shifts sheet back in the order the week actually runs, and stretch
    the filter over whatever is now there. Called after anything appends rows."""
    cols = columns_of(ws)
    width = ws.max_column
    rows = [[ws.cell(r, c).value for c in range(1, width + 1)]
            for r in range(2, ws.max_row + 1)]
    rows = [row for row in rows if any(s(v) for v in row)]
    day, start, name = cols["day"] - 1, cols["start"] - 1, cols["name"] - 1
    rows.sort(key=lambda row: (
        DAYS.index(s(row[day])) if s(row[day]) in DAYS else len(DAYS),
        minutes(row[start]),
        s(row[name]).lower(),
    ))
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r, c, value)
            if c in (cols["start"], cols["end"]):
                cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(width)}{max(len(rows) + 1, 2)}"
    return len(rows)


def minutes(value):
    """9:00 AM -> 540. Anything unreadable sorts to the end of its day."""
    text = s(value).upper().replace(".", "")
    try:
        clock, meridiem = text.split(" ")[0], text.split(" ")[-1]
        hour, minute = (clock.split(":") + ["0"])[:2]
        hour, minute = int(hour) % 12, int(minute)
        return hour * 60 + minute + (720 if meridiem.startswith("P") else 0)
    except (ValueError, IndexError):
        return 24 * 60


INSTRUCTIONS = """HOW TO EDIT THIS WORKBOOK

Two sheets carry the schedule: people and shifts. Everything else is optional.

One row per shift. Someone with three hours a week gets three rows on the
shifts sheet. Never put two names in one cell.

Both sheets have a filter row - click the arrow on any heading to narrow the
sheet to one name, one day, or one room. It does not change what students see.

- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
people - everyone who holds hours, listed once

  name       the exact name used on the shifts sheet. Must be unique.
             This is also what students see, so write it the way they should
             read it (Dr. Torres, not dtorres).
  role       faculty, gtf, or la. Pick from the dropdown.
             Anyone who tutors is an la, whether or not they also hold ENGR
             office hours.
  courses    what they can help with AT TUTORING in AV C141, separated by ;
             Blank means they do not tutor.
             This is the only place to edit it - every one of their AV C141
             shifts reads from this cell.
             ENGR office hours in AV C144 always advertises the courses in
             default_courses on the settings sheet, whatever is written here.
  languages  languages BESIDES ENGLISH they can help in, separated by ;
             Blank for most people. Becomes the Language filter and a line in
             the panel on the site.

  Yellow cells were guessed during the migration - please check them.

- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
shifts - one row per weekly hour

  name        pick from the dropdown (it reads the people sheet)
  day         Monday through Sunday. Days with no rows show as closed.
  start / end like 9:00 AM or 1:30 PM. Must land on the hour or half hour.
  location    WHERE THE HOUR HAPPENS. Fill this in on every row.
                AV C144            ENGR office hours
                AV C141            tutoring
                anywhere else      just type it - Scott Engineering, and so on.
                                   The site gives every room its own colour.
                a meeting link     paste it and the hour shows as Online.
              The first two are on the dropdown; you can type over it.
              Blank falls back to default_location on the settings sheet.

  Rows are sorted the way the week runs, but the site does not care about the
  order. Add new ones at the bottom if that is easier.

- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
exceptions - one-off changes to the normal week

  Cancel one shift:                name, date, start, end, type = cancelled
  Cancel a whole day for someone:  leave start and end blank
  Add extra hours:                 name, date, start, end, type = added

- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
settings - term dates, the announcement banner, and a few defaults

  default_courses    what an ENGR office hours shift advertises.
  default_location   the room a blank location falls back to.
  tutoring_room      the room that runs on the people sheet's course lists.
                     One edit moves the whole tutoring programme.
  subjects           groups the courses in the Get help with menu.
  course_order       the order that menu lists them in.
  announcement       a banner across the top of the site. Blank for none.

- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
CHECK YOUR WORK
Open the site with ?check on the end of the address. It lists anything it
could not read, row by row.

When you are done: save, then send the file to Ben (or save it back to
OneDrive)."""


def write_instructions(wb):
    name = "how to use this"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLORS[name]
    ws.column_dimensions["A"].width = 82
    for r, line in enumerate(INSTRUCTIONS.split("\n"), start=1):
        cell = ws.cell(r, 1, line or None)
        if r == 1 or (line and not line.startswith(" ") and line == line.upper() and len(line) > 3):
            cell.font = Font(bold=True)
    wb.move_sheet(ws, offset=len(wb.sheetnames))
    wb.active = 0


if __name__ == "__main__":
    main()
