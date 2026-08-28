#!/usr/bin/env python3
"""
Folds the AV C141 tutoring programme into data/office-hours.xlsx.

Two source grids, both kept in source-data/ (gitignored):

  Fall 2026 Classes and Tutors Grid.xlsx   tutor x course matrix, an X per course
  Fall 2026 Semifinal Schedule.xlsx        the rota, "Safwan A. (PH/MATH/ENGR)" per cell

Neither is a shape the site can read, and neither is the shape Ben maintains. This
script edits the published workbook *in place* rather than regenerating it, so the
yellow role flags, the comments and every hand edit made since the migration survive.

It is re-runnable: tutoring rows and tutoring course lists are rebuilt from scratch
each time, office-hours rows are never touched.

    python3 tools/add_tutoring.py

Anything it cannot resolve is printed and nothing is written.
"""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "source-data"
BOOK = ROOT / "data" / "office-hours.xlsx"
GRID = SOURCE / "Fall 2026 Classes and Tutors Grid.xlsx"
ROTA = SOURCE / "Fall 2026 Semifinal Schedule.xlsx"

PROGRAM = "tutoring"
ROOM = "AV C141"
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# The grid heads one column "PRECALC" with no course number under it. Precalculus is
# several MATH numbers at CSU and the grid does not say which, so it stays a plain
# label rather than a number we would be inventing.
PRECALC = "Precalculus"

# Spellings that differ between the tutoring grid and the office-hours workbook, which
# already committed to one of them. Same person; the workbook's spelling wins so the
# Person filter does not list them twice.
SAME_PERSON = {"Mia Shulze": "Mia Schulze"}

problems = []


def problem(message):
    problems.append(message)


def s(value):
    return "" if value is None else str(value).strip()


# ------------------------------------------------------------------ the matrix

def read_classes_grid():
    """{tutor name: [courses]} from the X matrix, plus the column order courses appear in."""
    ws = openpyxl.load_workbook(GRID, data_only=True)["Classes Grid"]

    columns = []          # (column index, [courses that column stands for])
    order = []            # every course, in the grid's own left-to-right order
    for col in range(1, ws.max_column + 1):
        subject = s(ws.cell(1, col).value)
        if not subject:
            continue
        number = s(ws.cell(2, col).value)
        if not number:
            courses = [PRECALC] if subject.upper().startswith("PRECALC") else [subject]
        else:
            # "111/112" is one column standing for two courses - a lecture and its lab.
            courses = [f"{subject} {part.strip()}" for part in number.split("/")]
        columns.append((col, courses))
        order.extend(c for c in courses if c not in order)

    tutors = {}
    for row in range(3, ws.max_row + 1):
        name = s(ws.cell(row, 5).value)
        if not name:
            continue
        name = SAME_PERSON.get(name, name)
        if name in tutors:
            problem(f"Classes Grid row {row}: {name} is listed twice.")
            continue
        courses = []
        for col, names in columns:
            if s(ws.cell(row, col).value).upper() == "X":
                courses.extend(names)
        tutors[name] = courses
    return tutors, order


# -------------------------------------------------------------------- the rota

TIME_RANGE = re.compile(r"^\s*(\d{1,2})(?::(\d{2}))?\s*-\s*(\d{1,2})(?::(\d{2}))?\s*$")
ENTRY = re.compile(r"^(.*?)\s*\(([^)]*)\)?\s*$")


def clock(hour, minute):
    """The rota is written on a 12-hour clock with no meridiem: "3:00 - 4:00".

    Every hour on it is afternoon or evening — the programme runs 3 PM to 9 PM —
    so all of them are PM, the closing 9 included. Reading 9 as morning silently
    produced nine shifts that ended before they started."""
    hour = int(hour)
    if not 3 <= hour <= 9:
        problem(f"Rota time \"{hour}\" is outside the 3-9 PM the programme runs.")
    return f"{hour % 12 or 12}:{minute or '00'} PM"


def read_rota():
    """[(short name, subjects, day, start, end)] plus the TBD slots, which have no name."""
    ws = openpyxl.load_workbook(ROTA, data_only=True)["FINAL"]

    # The sheet is two stacked tables - Monday-Thursday, then a Sunday one added
    # later - each under its own row of day names. Anything to the right of column J
    # is the tutor-letter legend and the bilingual note, not schedule.
    last_col = 10
    headers = []
    for row in range(1, ws.max_row + 1):
        found = {}
        for col in range(2, last_col + 1):
            value = s(ws.cell(row, col).value)
            if value in DAYS:
                found[col] = value
        if found:
            headers.append((row, found))
    if not headers:
        problem("FINAL sheet: found no row of day names.")
        return [], []

    bounds = [(r, headers[i + 1][0] if i + 1 < len(headers) else ws.max_row + 1)
              for i, (r, _) in enumerate(headers)]

    shifts, tbd = [], []
    for (header_row, columns), (_, stop) in zip(headers, bounds):
        start = end = None
        for row in range(header_row + 1, stop):
            label = s(ws.cell(row, 2).value)
            if label:
                m = TIME_RANGE.match(label)
                if not m:
                    problem(f"FINAL row {row}: cannot read the time \"{label}\".")
                    start = end = None
                    continue
                start, end = clock(m[1], m[2]), clock(m[3], m[4])
            for col, day in columns.items():
                text = s(ws.cell(row, col).value)
                if not text:
                    continue
                if start is None:
                    problem(f"FINAL {get_column_letter(col)}{row}: \"{text}\" sits above any time.")
                    continue
                if "tbd" in text.lower():
                    tbd.append((day, start, end, text.strip("* ")))
                    continue
                m = ENTRY.match(text)
                if not m:
                    problem(f"FINAL {get_column_letter(col)}{row}: cannot read \"{text}\".")
                    continue
                subjects = [p.strip().upper() for p in m[2].split("/") if p.strip()]
                shifts.append((m[1].strip(), subjects, day, start, end))
    return shifts, tbd


def read_languages():
    """The "Bilingual Tutors!" note beside the Sunday table: short name -> language."""
    ws = openpyxl.load_workbook(ROTA, data_only=True)["FINAL"]
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if "bilingual" not in s(ws.cell(row, col).value).lower():
                continue
            out = {}
            for r in range(row + 1, ws.max_row + 1):
                text = s(ws.cell(r, col).value)
                if not text:
                    break
                name, _, language = text.partition("-")
                if language:
                    out[name.strip()] = language.strip()
            return out
    return {}


# ------------------------------------------------------------- name resolution

def resolve(short, tutors):
    """"Safwan A." -> "Safwan Ahmad".

    The initial is matched against every part of the surname, not just the first:
    Joseph Quiroz Hernandez is "Joseph H." on the rota, and Jordy Medina Valverde
    is "Jordy V.". Ambiguity is reported rather than guessed at."""
    first, _, initial = short.replace(".", "").strip().partition(" ")
    first, initial = first.lower(), initial.strip().lower()
    hits = []
    for name in tutors:
        parts = name.lower().split()
        if parts[0] != first:
            continue
        if initial and not any(p.startswith(initial) for p in parts[1:]):
            continue
        hits.append(name)
    if len(hits) == 1:
        return hits[0]
    problem(f'Rota name "{short}" matches {hits or "nobody"} on the Classes Grid.'
            if hits else f'Rota name "{short}" matches nobody on the Classes Grid.')
    return None


def subject_of(course):
    head = re.match(r"^([A-Za-z]+)", course)
    return head[1].upper() if head else course.upper()


def subjects_of(courses):
    """The course list reduced to the subject codes the rota writes in brackets,
    so the two sources can be checked against each other."""
    out = set()
    for course in courses:
        out.add("MATH" if course == PRECALC else subject_of(course))
    return out


# ------------------------------------------------------------------- the write

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True)


def header_cell(ws, col, text, width):
    cell = ws.cell(row=1, column=col, value=text)
    cell.fill, cell.font = HEAD_FILL, HEAD_FONT
    cell.alignment = Alignment(vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def column_of(ws, name):
    for col in range(1, ws.max_column + 1):
        if s(ws.cell(1, col).value).lower() == name:
            return col
    return None


def dropdown(ws, formula, message, cells):
    """Replaces any existing validation over the same cells - openpyxl keeps the old
    one on load, and two list validations on one range make Excel complain."""
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation if str(dv.sqref) != cells
    ]
    dv = DataValidation(type="list", formula1=formula, allowBlank=True)
    dv.showErrorMessage = True
    dv.errorTitle = "Not one of the choices"
    dv.error = message
    ws.add_data_validation(dv)
    dv.add(cells)


def set_setting(ws, key, value):
    for row in range(2, ws.max_row + 1):
        if s(ws.cell(row, 1).value).lower() == key:
            ws.cell(row, 2, value).number_format = "@"
            return
    row = ws.max_row + 1
    ws.cell(row, 1, key)
    ws.cell(row, 2, value).number_format = "@"


def main():
    for path in (GRID, ROTA, BOOK):
        if not path.exists():
            sys.exit(f"missing: {path}")

    tutors, course_order = read_classes_grid()
    rota, tbd = read_rota()
    languages = read_languages()

    # Resolve every short name once, and check the rota's bracketed subjects against
    # the matrix while we are here - they are two people's edits of the same fact.
    resolved = []
    for short, subjects, day, start, end in rota:
        full = resolve(short, tutors)
        if not full:
            continue
        known = subjects_of(tutors[full])
        missing = [x for x in subjects if x not in known]
        if missing:
            problem(f"{full} is scheduled as {'/'.join(subjects)} but the Classes Grid "
                    f"has no {', '.join(missing)} for them.")
        resolved.append((full, day, start, end))

    for short in languages:
        resolve(short, tutors)

    if problems:
        print("Nothing written. Fix these first:\n")
        for p in problems:
            print(f"  - {p}")
        sys.exit(1)

    wb = openpyxl.load_workbook(BOOK)
    people, shifts, settings, readme = wb["people"], wb["shifts"], wb["settings"], wb["how to use this"]

    # -- people: a second course list, for what they tutor -----------------
    col_courses_tutoring = column_of(people, "courses_tutoring") or people.max_column + 1
    header_cell(people, col_courses_tutoring, "courses_tutoring", 44)
    # The bilingual note gets a column of its own rather than a sentence in
    # notes. Notes is prose nobody can filter on, and writing there clobbered
    # whatever a human had put in the cell.
    col_languages = column_of(people, "languages") or people.max_column + 1
    header_cell(people, col_languages, "languages", 18)
    col_name, col_role = column_of(people, "name"), column_of(people, "role")

    rows = {s(people.cell(r, col_name).value): r
            for r in range(2, people.max_row + 1) if s(people.cell(r, col_name).value)}

    language_by_full = {}
    for short, language in languages.items():
        language_by_full[resolve(short, tutors)] = language

    added = []
    for name, courses in tutors.items():
        row = rows.get(name)
        if row is None:
            row = people.max_row + 1
            people.cell(row, col_name, name)
            # Someone off the tutoring grid who holds no ENGR office hours. They
            # are still a Learning Assistant: the separate tutor role was folded
            # into la, because to a student it named the same kind of help.
            people.cell(row, col_role, "la")
            rows[name] = row
            added.append(name)
        people.cell(row, col_courses_tutoring, "; ".join(courses))

    # Languages are rebuilt from the rota every run, for everyone: someone taken
    # off the bilingual list should stop being advertised as speaking it.
    for name, row in rows.items():
        people.cell(row, col_languages, language_by_full.get(name) or None)

    # Clear stale tutoring lists for anyone no longer on the grid.
    for name, row in rows.items():
        if name not in tutors:
            people.cell(row, col_courses_tutoring, None)

    dropdown(people, '"faculty,gtf,la"',
             "Role must be faculty, gtf, or la.",
             f"{get_column_letter(col_role)}2:{get_column_letter(col_role)}400")

    # -- shifts: which programme a row belongs to --------------------------
    col_program = column_of(shifts, "program") or shifts.max_column + 1
    header_cell(shifts, col_program, "program", 14)
    col_shift_name = column_of(shifts, "name")
    col_day, col_start, col_end = column_of(shifts, "day"), column_of(shifts, "start"), column_of(shifts, "end")

    # Drop the tutoring rows from any previous run, keep every office-hours row.
    for row in range(shifts.max_row, 1, -1):
        if s(shifts.cell(row, col_program).value).lower() == PROGRAM:
            shifts.delete_rows(row)

    first_new = shifts.max_row + 1
    for i, (name, day, start, end) in enumerate(sorted(
            resolved, key=lambda x: (DAYS.index(x[1]), x[2].zfill(8), x[0]))):
        row = first_new + i
        shifts.cell(row, col_shift_name, name)
        shifts.cell(row, col_day, day)
        shifts.cell(row, col_start, start).alignment = Alignment(horizontal="left")
        shifts.cell(row, col_end, end).alignment = Alignment(horizontal="left")
        shifts.cell(row, col_program, PROGRAM)
        # location and courses stay blank: the room comes from
        # default_location_tutoring, the courses from courses_tutoring.

    letter = get_column_letter(col_program)
    dropdown(shifts, f'"office hours,{PROGRAM}"',
             "Use office hours or tutoring. Blank means office hours.",
             f"{letter}2:{letter}400")
    dropdown(shifts, f'"{",".join(DAYS)}"', "Pick a day of the week.",
             f"{get_column_letter(col_day)}2:{get_column_letter(col_day)}400")

    # -- settings ----------------------------------------------------------
    set_setting(settings, "default_program", "office hours")
    set_setting(settings, "program_order", "office hours; tutoring")
    set_setting(settings, f"default_location_{PROGRAM}", ROOM)
    set_setting(settings, "subjects",
                "ENGR -> Engineering; MATH -> Math; Precalculus -> Math; CHEM -> Chemistry; "
                "PH -> Physics; CIVE -> Civil Engineering; CS -> Computer Science; LIFE -> Biology")
    # ENGR first - this is the ENGR site - then the grid's own subject order, but
    # numerically within each subject: the grid happens to head CIVE 261 before
    # CIVE 260, which in a menu just looks like a mistake.
    ordered = ["ENGR 111", "ENGR 114", "ENGR 123"]
    for subject in dict.fromkeys(subject_of(c) for c in course_order):
        ordered += sorted((c for c in course_order
                           if subject_of(c) == subject and c not in ordered),
                          key=lambda c: c.split(" ")[-1])
    set_setting(settings, "course_order", "; ".join(ordered))

    # -- the tab that explains all this to whoever edits next --------------
    rewrite_readme(readme)

    if tbd:
        note = "; ".join(f"{day} {start}" for day, start, _, _ in tbd)
        print(f"note: {len(tbd)} slot(s) still marked TBD on the rota and left off the "
              f"site until a name is filled in — {note}")

    wb.save(BOOK)
    print(f"{len(tutors)} tutors, {len(resolved)} tutoring shifts written to {BOOK.relative_to(ROOT)}")
    print(f"new to the workbook: {', '.join(added) or 'nobody'}")


README = [
    ("HOW TO EDIT THIS WORKBOOK", True),
    ("", False),
    ("One row per shift. If someone holds three office hours a week, they get three rows", False),
    ("on the shifts sheet. Never put two names in one cell.", False),
    ("", False),
    ("The workbook covers two things: ENGR office hours, and tutoring in AV C141.", False),
    ("A person can do both - they are listed once and get a row per shift either way.", False),
    ("", False),
    ("people sheet - everyone who holds hours, listed once", True),
    ("  name          the exact name used on the shifts sheet. Must be unique.", False),
    ("  display_name  what students see, if different (e.g. Dr. Torres). Blank = same as name.", False),
    ("  role          faculty, gtf, or la. Pick from the dropdown.", False),
    ("                Everyone who tutors is an la, whether or not they also hold", False),
    ("                ENGR office hours.", False),
    ("  courses       what they help with at ENGR OFFICE HOURS.", False),
    ("                Leave BLANK for the default (ENGR 111 and ENGR 114).", False),
    ("                If you fill it in, list ALL courses they cover, separated by ;", False),
    ("                Anyone covering ENGR 111 is counted for ENGR 123 too - see the", False),
    ("                course_implies row on the settings sheet.", False),
    ("  courses_tutoring   what they can help with AT TUTORING. Separated by ;", False),
    ("                Blank means they do not tutor. This is the only place to edit it -", False),
    ("                every one of their tutoring shifts reads from this cell.", False),
    ("  languages     languages BESIDES ENGLISH they can help in, separated by ;", False),
    ("                Blank for most people. Becomes the Language filter and a", False),
    ("                line in the panel. Rebuilt from the rota's bilingual list.", False),
    ("  email, notes  optional. Both are shown to students, so keep notes short and useful.", False),
    ("", False),
    ("  Yellow cells were guessed during the migration - please check them.", False),
    ("", False),
    ("shifts sheet - one row per weekly hour", True),
    ("  name          pick from the dropdown (it reads the people sheet)", False),
    ("  day           Monday through Sunday. Days with no rows show as closed.", False),
    ("  start / end   like 9:00 AM or 1:30 PM. Must land on the hour or half hour.", False),
    ("  program       office hours or tutoring. Blank = office hours.", False),
    ("                Tutoring rows take their room and their courses from tutoring:", False),
    ("                leave location and courses blank unless this one row differs.", False),
    ("  mode          in-person or online. Blank = in-person.", False),
    ("  location      blank = the default room for the programme (see settings).", False),
    ("                Fill it in for anywhere else - Scott Engineering, etc.", False),
    ("                The site gives every room its own colour automatically.", False),
    ("                For online, put the meeting link here.", False),
    ("  courses       blank = whatever the people sheet says for that programme.", False),
    ("  active        put no to hide a row for a while without deleting it.", False),
    ("", False),
    ("exceptions sheet - one-off changes to the normal week", True),
    ("  Cancel one shift:  name, date, start, end, type = cancelled", False),
    ("  Cancel a whole day for someone: leave start and end blank", False),
    ("  Add extra hours:   name, date, start, end, type = added", False),
    ("", False),
    ("settings sheet - term dates, default rooms, and the announcement banner", True),
    ("  default_location_tutoring   the tutoring room. One edit moves the whole", False),
    ("                              programme.", False),
    ("  subjects                    groups the courses in the Get help with menu.", False),
    ("  course_order                the order that menu lists them in.", False),
    ("", False),
    ("When you are done: save, then send the file to Ben (or save it back to OneDrive).", False),
]


def rewrite_readme(ws):
    for row in range(ws.max_row, 0, -1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None
    ws.column_dimensions["A"].width = 110
    for r, (text, bold) in enumerate(README, start=1):
        cell = ws.cell(row=r, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)


if __name__ == "__main__":
    main()
