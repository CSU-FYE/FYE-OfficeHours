#!/usr/bin/env python3
"""One-time migration: the two original grid workbooks -> the standardized workbook.

You should never need to run this. It was used once to produce data/office-hours.xlsx
from the Fall 2026 source schedules. It is kept so the conversion is auditable and
re-runnable if the originals are corrected.

    python3 tools/make_workbook.py

Writes data/office-hours.xlsx and MIGRATION-NOTES.md.
"""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "source-data"
OUT = ROOT / "data" / "office-hours.xlsx"
NOTES = ROOT / "MIGRATION-NOTES.md"

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
# Where the old grids said each group met. Used only to spot per-cell overrides:
# a room matching the group's own default is dropped so the settings sheet stays
# the single place a room is changed.
ROOM_FACULTY = "AV C147"
ROOM_LA = "AV C144"
# Where everyone meets now. Faculty and GTFs moved out of C147 into C144, so the
# blanks written above resolve to this.
ROOM_DEFAULT = "AV C144"

# Roles we could not read off the source with certainty. Every name here is
# reported in MIGRATION-NOTES.md for review.
GUESSED_ROLES = {
    "Dr. Dan": "faculty",
    "Dr. Harvey": "faculty",
    "Dr. Yume": "faculty",
    "DrT": "faculty",
    "Prof Scheller": "faculty",
    "Chloe": "gtf",
    "Taylor": "gtf",
    "Ben": "gtf",
    "Winnie": "gtf",
    "Ojo": "gtf",
    "Sumaiya": "gtf",
}

flags = []


def flag(msg):
    flags.append(msg)


# ---------------------------------------------------------------- parsing ---

TIME_RE = re.compile(
    r"^\s*(\d{1,2})(?::(\d{2}))?\s*(am|pm)?\s*(?:to|-|–)\s*(\d{1,2})(?::(\d{2}))?\s*(am|pm)?\s*$",
    re.I,
)


def to_24h(hour, minute, meridiem):
    hour, minute = int(hour), int(minute or 0)
    if meridiem:
        meridiem = meridiem.lower()
        if meridiem == "pm" and hour != 12:
            hour += 12
        if meridiem == "am" and hour == 12:
            hour = 0
    return hour, minute


def parse_time_range(label):
    """'9 am to 10 am' / '3pm to 4pm' / '11 am to 12 pm ' -> ('9:00 AM', '10:00 AM')."""
    if not label:
        return None
    m = TIME_RE.match(str(label).replace(" ", " "))
    if not m:
        return None
    sh, sm, smer, eh, em, emer = m.groups()
    # A missing meridiem on the start borrows the end's, then corrects for a
    # range that would run backwards ('11 am to 12 pm' must not become 11pm).
    e_hour, e_min = to_24h(eh, em, emer)
    s_hour, s_min = to_24h(sh, sm, smer or emer)
    if s_hour * 60 + s_min >= e_hour * 60 + e_min:
        s_hour, s_min = to_24h(sh, sm, "am" if (smer or emer or "").lower() == "pm" else "pm")
    return fmt(s_hour, s_min), fmt(e_hour, e_min)


def fmt(hour, minute):
    mer = "AM" if hour < 12 else "PM"
    h12 = hour % 12 or 12
    return f"{h12}:{minute:02d} {mer}"


ROOM_RE = re.compile(r"\(?\s*(AV\s*C?\s*\d{3})\s*\)?", re.I)
SPLIT_RE = re.compile(r"\s*(?:&|,|\band\b|\n|\r)\s*", re.I)
TRAILING_ROLE_RE = re.compile(r"\s*\(\s*GTAs?\s*\)\s*$", re.I)


def normalize_room(raw):
    """'AV147' / 'AV 147' / 'AVC147' -> 'AV C147'."""
    digits = re.sub(r"\D", "", raw)
    return f"AV C{digits}"


def parse_cell(text):
    """A grid cell -> [(name, room_override_or_None), ...].

    Handles 'Chloe (AV147) and DrT', 'Prof Scheller (AV 147) & Dr. Yume',
    'Sumaiya - AV147', and newline-separated names.
    """
    if not text:
        return []
    out = []
    for chunk in SPLIT_RE.split(str(text)):
        chunk = chunk.strip().strip("-").strip()
        if not chunk:
            continue
        room = None
        m = ROOM_RE.search(chunk)
        if m:
            room = normalize_room(m.group(1))
            chunk = ROOM_RE.sub("", chunk).strip().strip("-").strip()
        chunk = TRAILING_ROLE_RE.sub("", chunk).strip()
        chunk = re.sub(r"\s{2,}", " ", chunk)
        if chunk:
            out.append((chunk, room))
    return out


def read_grid(path, sheet, time_col, first_day_col, header_row):
    """Read a time-rows x day-columns grid into shift tuples."""
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    shifts = []
    for row in range(header_row + 1, ws.max_row + 1):
        span = parse_time_range(ws.cell(row=row, column=time_col).value)
        if not span:
            continue
        start, end = span
        for i, day in enumerate(DAYS):
            cell = ws.cell(row=row, column=first_day_col + i).value
            for name, room in parse_cell(cell):
                shifts.append(
                    {
                        "name": name,
                        "day": day,
                        "start": start,
                        "end": end,
                        "room": room,
                        "src": f"{Path(path).name} · {sheet} · {get_column_letter(first_day_col + i)}{row}",
                    }
                )
    return shifts


# ------------------------------------------------------------------ build ---

def main():
    faculty_src = SRC / "FA26_Faculty_GTF Office Hours.xlsx"
    la_src = SRC / "Fall 2026 LA Office Hours.xlsx"
    for p in (faculty_src, la_src):
        if not p.exists():
            sys.exit(f"missing source workbook: {p}")

    faculty_shifts = read_grid(faculty_src, "Office Hours", 1, 2, 2)
    la_shifts = read_grid(la_src, "Learning Assistants", 2, 3, 4)

    people = {}
    for s in faculty_shifts:
        people.setdefault(s["name"], {"role": None, "src": s["src"], "sheet": "faculty/GTF"})
    for s in la_shifts:
        people.setdefault(s["name"], {"role": "la", "src": s["src"], "sheet": "LA"})

    for name, info in people.items():
        if info["role"] is None:
            info["role"] = GUESSED_ROLES.get(name, "gtf")
            info["guessed"] = True
            flag(
                f"**{name}** — role guessed as `{info['role']}`. Appears only as "
                f"`{name}` on the Faculty & GTF grid (first seen {info['src']}). "
                f"Confirm the full name and whether they are faculty or a GTF."
            )

    shifts = []
    for s in faculty_shifts + la_shifts:
        role = people[s["name"]]["role"]
        default_room = ROOM_LA if role == "la" else ROOM_FACULTY
        room = s["room"] or default_room
        shifts.append(
            {
                "name": s["name"],
                "day": s["day"],
                "start": s["start"],
                "end": s["end"],
                # Leave location blank when it matches the role default, so the
                # settings sheet stays the single place a room is changed.
                "location": "" if room == default_room else room,
                "src": s["src"],
            }
        )

    shifts.sort(key=lambda s: (DAYS.index(s["day"]), s["start"], s["name"]))

    write_workbook(sorted(people.items()), shifts)
    write_notes(people, shifts)
    print(f"wrote {OUT.relative_to(ROOT)}  ({len(people)} people, {len(shifts)} shifts)")
    print(f"wrote {NOTES.relative_to(ROOT)}  ({len(flags)} items to review)")


def dropdown(formula, message):
    """A list validation that blocks bad typing but still allows an empty cell —
    most rows on these sheets are legitimately blank until someone fills them in."""
    dv = DataValidation(type="list", formula1=formula, allowBlank=True)
    dv.showErrorMessage = True
    dv.errorTitle = "Not one of the choices"
    dv.error = message
    return dv


HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True)


def style_header(ws, headers, widths, note_row=None):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_workbook(people_items, shifts):
    wb = openpyxl.Workbook()

    # -- people ---------------------------------------------------------
    ws = wb.active
    ws.title = "people"
    style_header(
        ws,
        ["name", "display_name", "role", "courses", "email", "notes"],
        [26, 22, 12, 30, 26, 46],
    )
    for r, (name, info) in enumerate(people_items, start=2):
        ws.cell(row=r, column=1, value=name)
        role_cell = ws.cell(row=r, column=3, value=info["role"])
        if info.get("guessed"):
            # Flagged in the sheet, not in `notes` — that column is shown to students.
            role_cell.fill = PatternFill("solid", fgColor="FFF2CC")
            role_cell.comment = Comment(
                "Guessed during the migration from the old schedules, because the grid "
                "only gave a first name or a title. Please confirm and clear the "
                "highlight. See MIGRATION-NOTES.md.",
                "migration",
            )

    dv_role = dropdown('"faculty,gtf,la"', "Role must be faculty, gtf, or la.")
    ws.add_data_validation(dv_role)
    dv_role.add("C2:C400")

    # -- shifts ---------------------------------------------------------
    sh = wb.create_sheet("shifts")
    style_header(
        sh,
        ["name", "day", "start", "end", "mode", "location", "courses", "active", "notes"],
        [26, 12, 11, 11, 12, 20, 26, 9, 40],
    )
    for r, s in enumerate(shifts, start=2):
        sh.cell(row=r, column=1, value=s["name"])
        sh.cell(row=r, column=2, value=s["day"])
        sh.cell(row=r, column=3, value=s["start"]).alignment = Alignment(horizontal="left")
        sh.cell(row=r, column=4, value=s["end"]).alignment = Alignment(horizontal="left")
        sh.cell(row=r, column=6, value=s["location"])
    # Times stay text so Excel cannot reformat them into something ambiguous.
    # Set on the column, not per cell — that would leave 400 empty styled cells
    # behind and send Ctrl+End to the bottom of the sheet.
    for col in ("C", "D"):
        sh.column_dimensions[col].number_format = "@"

    dv_day = dropdown('"Monday,Tuesday,Wednesday,Thursday,Friday"', "Office hours are Monday to Friday.")
    dv_mode = dropdown('"in-person,online"', "Use in-person or online. Blank means in-person.")
    dv_active = dropdown('"yes,no"', "Use yes or no. Blank means yes.")
    dv_name = dropdown("PeopleNames", "Pick a name from the people sheet. Add them there first if they are new.")
    for dv, rng in ((dv_day, "B2:B400"), (dv_mode, "E2:E400"), (dv_active, "H2:H400"), (dv_name, "A2:A400")):
        sh.add_data_validation(dv)
        dv.add(rng)

    # -- exceptions -----------------------------------------------------
    ex = wb.create_sheet("exceptions")
    style_header(
        ex,
        ["name", "date", "start", "end", "type", "mode", "location", "note"],
        [26, 14, 11, 11, 12, 12, 20, 46],
    )
    ex["A2"] = ""
    dv_type = dropdown('"cancelled,added"', "Use cancelled or added.")
    dv_ex_name = dropdown("PeopleNames", "Pick a name from the people sheet.")
    for dv, rng in ((dv_type, "E2:E400"), (dv_ex_name, "A2:A400")):
        ex.add_data_validation(dv)
        dv.add(rng)
    for col in ("B", "C", "D"):
        ex.column_dimensions[col].number_format = "@"

    # -- settings -------------------------------------------------------
    st = wb.create_sheet("settings")
    style_header(st, ["key", "value"], [30, 60])
    settings = [
        ("term_name", "Fall 2026"),
        ("term_start", "2026-08-24"),
        ("term_end", "2026-12-11"),
        ("day_start", "9:00 AM"),
        ("day_end", "9:00 PM"),
        ("default_location_faculty", ROOM_DEFAULT),
        ("default_location_gtf", ROOM_DEFAULT),
        ("default_location_la", ROOM_DEFAULT),
        ("room_order", "AV C144; AV C141; Scott Engineering"),
        ("default_courses", "ENGR 111; ENGR 114"),
        ("timezone", "America/Denver"),
        ("announcement", ""),
    ]
    for r, (k, v) in enumerate(settings, start=2):
        st.cell(row=r, column=1, value=k)
        st.cell(row=r, column=2, value=v).number_format = "@"

    # -- readme ---------------------------------------------------------
    rd = wb.create_sheet("how to use this")
    rd.column_dimensions["A"].width = 110
    lines = [
        ("HOW TO EDIT THIS WORKBOOK", True),
        ("", False),
        ("One row per shift. If someone holds three office hours a week, they get three rows", False),
        ("on the shifts sheet. Never put two names in one cell.", False),
        ("", False),
        ("people sheet - everyone who holds office hours, listed once", True),
        ("  name          the exact name used on the shifts sheet. Must be unique.", False),
        ("  display_name  what students see, if different (e.g. Dr. Torres). Blank = same as name.", False),
        ("  role          faculty, gtf, or la. Pick from the dropdown.", False),
        ("  courses       leave BLANK if they help with the default (ENGR 111 and ENGR 114).", False),
        ("                If you fill it in, list ALL courses they cover, separated by ;", False),
        ("  email, notes  optional. Both are shown to students, so keep notes short and useful.", False),
        ("", False),
        ("  Yellow cells were guessed during the migration - please check them.", False),
        ("", False),
        ("shifts sheet - one row per weekly office hour", True),
        ("  name          pick from the dropdown (it reads the people sheet)", False),
        ("  day           Monday through Friday", False),
        ("  start / end   like 9:00 AM or 1:30 PM. Must land on the hour or half hour.", False),
        ("  mode          in-person or online. Blank = in-person.", False),
        ("  location      blank = the default room for their role (see settings).", False),
        ("                Fill it in for anywhere else - AV C141, Scott Engineering, etc.", False),
        ("                The site gives every room its own colour automatically.", False),
        ("                For online, put the meeting link here.", False),
        ("  active        put no to hide a row for a while without deleting it.", False),
        ("", False),
        ("exceptions sheet - one-off changes to the normal week", True),
        ("  Cancel one shift:  name, date, start, end, type = cancelled", False),
        ("  Cancel a whole day for someone: leave start and end blank", False),
        ("  Add extra hours:   name, date, start, end, type = added", False),
        ("", False),
        ("settings sheet - term dates, default rooms, and the announcement banner", True),
        ("", False),
        ("When you are done: save, then send the file to Ben (or save it back to OneDrive).", False),
    ]
    for r, (text, bold) in enumerate(lines, start=1):
        c = rd.cell(row=r, column=1, value=text)
        if bold:
            c.font = Font(bold=True)

    wb.defined_names.add(DefinedName("PeopleNames", attr_text="people!$A$2:$A$400"))
    wb.active = 0
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)


def write_notes(people, shifts):
    by_person = {}
    for s in shifts:
        by_person.setdefault(s["name"], []).append(s)

    lines = [
        "# Migration notes",
        "",
        "How the two original grid workbooks in `source-data/` were converted into the",
        "standardized `data/office-hours.xlsx`. **Everything the conversion had to guess is",
        "listed here.** Fix anything wrong directly in the workbook — this file is a record,",
        "not an input.",
        "",
        "## Sources used",
        "",
        "| Source | Sheet | Used |",
        "|---|---|---|",
        "| `FA26_Faculty_GTF Office Hours.xlsx` | `Office Hours` | yes |",
        "| `Fall 2026 LA Office Hours.xlsx` | `Learning Assistants` | yes |",
        "| `Fall 2026 LA Office Hours.xlsx` | `Professor & GTFs` | no — identical to the faculty workbook |",
        "| `Fall 2026 LA Office Hours.xlsx` | `Master Schedule` | **no — see below** |",
        "| `Fall 2026 LA Office Hours.xlsx` | `Sheet2` | no — raw availability survey, an input not a schedule |",
        "| `Fall 2026 LA Office Hours.xlsx` | `Cancellations` | no — header row only, empty |",
        "",
        "## Please confirm",
        "",
    ]
    for i, f in enumerate(flags, start=1):
        lines.append(f"{i}. {f}")
    lines += [
        "",
        "### The `Master Schedule` tab was left out",
        "",
        "Its names mostly do not appear on the `Learning Assistants` tab — AJ Tripp, Max Goodrich,",
        "Maddie Potter, Trevor Ridge, Ethan Kramer and others appear nowhere else — and it stops at",
        "5pm. It reads like a previous term. **If it is current, say so and it can be merged in.**",
        "",
        "It also contains two notes to you: *\"Please put GTA beside your name if you are one\"* and",
        "*\"Chloe Brekhus - LA or GTA?\"* — that second question is still open, and matters for the",
        "role filter.",
        "",
        "### Names that are probably the same person",
        "",
        "The conversion did **not** merge these, because a wrong merge invents a schedule that",
        "nobody holds. If you confirm a pair, just rename in the workbook.",
        "",
        "| On the schedule | Possibly | Seen in |",
        "|---|---|---|",
        "| `DrT` | `Dr. Torres` | `Master Schedule` |",
        "| `Chloe` | `Chloe Brekhus` | `Master Schedule` (flagged there as \"LA or GTA?\") |",
        "| `Prof Scheller` | `Dylan Scheller (GTA)` | `Master Schedule` — if so the role is **gtf**, not faculty |",
        "| `Taylor` | `Nell Taylor` | `Master Schedule` |",
        "",
        "`Dr. Dan`, `Dr. Harvey`, `Dr. Yume`, `Winnie`, `Ojo`, `Sumaiya` and `Ben` appear only as",
        "these short forms. Full names would let students search for them by surname.",
        "",
        "## What was normalized automatically",
        "",
        "- `3pm to 4pm` and `11 am to 12 pm ` (trailing space) → consistent `3:00 PM` / `11:00 AM` times",
        "- `Chloe (AV147) and DrT` → two separate rows; the room moved into its own column",
        "- `Prof Scheller (AV 147) & Dr. Yume`, and names split across two lines in one cell → separate rows",
        "- `Sumaiya - AV147` → name and room separated",
        "- `AV147`, `AV 147` → `AV C147`",
        "- `(GTA)` suffixes stripped from names and turned into the `role` column",
        "- Rooms matching the role default are left blank, so a room change is a one-line edit in `settings`",
        "- Everyone was set to **in-person**. The `Online Office hours` section of the LA sheet was empty.",
        "",
        "## Result",
        "",
        f"- **{len(people)} people** — {sum(1 for p in people.values() if p['role'] == 'la')} LAs, "
        f"{sum(1 for p in people.values() if p['role'] == 'gtf')} GTFs, "
        f"{sum(1 for p in people.values() if p['role'] == 'faculty')} faculty",
        f"- **{len(shifts)} shifts**, {sum(1 for s in shifts if s['location'])} with a room other than their role default",
        "",
        "### Every shift, and the cell it came from",
        "",
        "| Person | Day | Start | End | From |",
        "|---|---|---|---|---|",
    ]
    for s in shifts:
        lines.append(f"| {s['name']} | {s['day']} | {s['start']} | {s['end']} | `{s['src']}` |")
    NOTES.write_text("\n".join(lines) + "\n")


if __name__ == "__main__":
    main()
