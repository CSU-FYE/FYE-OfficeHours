#!/usr/bin/env python3
"""
Folds a fresh Faculty & GTF grid into data/office-hours.xlsx.

    source-data/FA26_Faculty_GTF Office Hours.xlsx    time-rows x day-columns grid

That grid is the shape the department maintains; it is not a shape the site can read,
and re-running make_workbook.py to convert it would throw away every hand edit made
since the migration. So this script edits the published workbook *in place*, the same
way add_tutoring.py does for AV C141.

It is re-runnable: the office-hours rows of everyone named on the grid are rebuilt
from scratch each time. LA rows, tutoring rows, roles, notes, emails and the yellow
role flags are never touched.

    python3 tools/update_faculty_hours.py          # show the diff and write
    python3 tools/update_faculty_hours.py --dry-run

Two things it will not do quietly:

  * A grid name that is new to the workbook is added to `people` with a *guessed*
    role, highlighted yellow with a comment, exactly like the migration's guesses.
  * Someone with faculty/GTF hours in the workbook who has vanished from the grid
    keeps their rows and is reported. A name disappearing is far more often a grid
    edited in a hurry than a person who stopped holding hours.
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
# The grid parser already exists and is already the audited one - importing it beats a
# second copy that could drift. make_workbook only runs under __main__, so this is safe.
from make_workbook import DAYS, GUESSED_ROLES, CONFIRMED_ROLES, read_grid

ROOT = Path(__file__).resolve().parent.parent
BOOK = ROOT / "data" / "office-hours.xlsx"
GRID = ROOT / "source-data" / "FA26_Faculty_GTF Office Hours.xlsx"
SHEET = "Office Hours"

# The grid still writes "(AV147)" beside half the names, left over from when faculty and
# GTFs met in C147. They moved to C144 and the office confirmed it, so the room text is
# read and dropped: location stays blank and resolves through default_location_<role>.
IGNORE_GRID_ROOMS = True

ROLE_GUESS_COMMENT = (
    "Guessed when this name arrived on the Faculty & GTF grid, which only gave a first "
    "name or a title. Please confirm and clear the highlight. See HANDOFF.md."
)


def s(value):
    return "" if value is None else str(value).strip()


def minutes(clock):
    """'1:00 PM' -> 780. Sorting on the text alone puts 9 AM after 10 AM."""
    body, meridiem = clock.rsplit(" ", 1)
    hour, minute = (int(x) for x in body.split(":"))
    hour = hour % 12 + (12 if meridiem.upper() == "PM" else 0)
    return hour * 60 + minute


def column_of(ws, name):
    for c in range(1, ws.max_column + 1):
        if s(ws.cell(1, c).value).lower() == name:
            return c
    sys.exit(f"the {ws.title!r} sheet has no {name!r} column")


def key(day, start, name):
    return (DAYS.index(day), minutes(start), name.lower())


def main():
    dry_run = "--dry-run" in sys.argv[1:]
    for path in (GRID, BOOK):
        if not path.exists():
            sys.exit(f"missing: {path}")

    grid = read_grid(GRID, SHEET, 1, 2, 2)
    if not grid:
        sys.exit(f"read no shifts at all from {GRID.name} - has the layout changed?")

    wanted = sorted({(g["name"], g["day"], g["start"], g["end"]) for g in grid},
                    key=lambda x: key(x[1], x[2], x[0]))
    grid_names = {g["name"] for g in grid}
    rooms = sorted({g["room"] for g in grid if g["room"]})

    wb = openpyxl.load_workbook(BOOK)
    people, shifts = wb["people"], wb["shifts"]

    p_name, p_role = column_of(people, "name"), column_of(people, "role")
    known = {s(people.cell(r, p_name).value): r
             for r in range(2, people.max_row + 1) if s(people.cell(r, p_name).value)}

    col = {n: column_of(shifts, n) for n in ("name", "day", "start", "end", "program")}

    def is_office_hours(row):
        return s(shifts.cell(row, col["program"]).value).lower() != "tutoring"

    def read(row):
        return (s(shifts.cell(row, col["name"]).value),
                s(shifts.cell(row, col["day"]).value),
                s(shifts.cell(row, col["start"]).value),
                s(shifts.cell(row, col["end"]).value))

    # What is there now, for the diff and for the "vanished from the grid" check.
    before = [read(r) for r in range(2, shifts.max_row + 1) if is_office_hours(r)]
    mine = {row for row in before if row[0] in grid_names}
    # A shift row can name someone the people sheet has never heard of; that is ?check=1's
    # problem to report, not a reason for this script to fall over.
    def role_of(name):
        row = known.get(name)
        return s(people.cell(row, p_role).value).lower() if row else ""

    dropped = sorted({row[0] for row in before
                      if row[0] not in grid_names and role_of(row[0]) in ("faculty", "gtf")})

    added = [w for w in wanted if w not in mine]
    removed = sorted(mine - set(wanted), key=lambda x: key(x[1], x[2], x[0]))
    new_people = [n for n in sorted(grid_names) if n not in known]

    print(f"{GRID.name}: {len(wanted)} shifts, {len(grid_names)} people")
    for label, rows in (("removed", removed), ("added", added)):
        for name, day, start, end in rows:
            print(f"  {'-' if label == 'removed' else '+'} {name}: {day} {start}–{end}")
    if not added and not removed:
        print("  (no change)")
    for name in new_people:
        print(f"  + {name} is new to the workbook")
    if rooms and IGNORE_GRID_ROOMS:
        print(f"  note: ignored room text on the grid ({', '.join(rooms)}); the room comes "
              f"from settings")
    for name in dropped:
        print(f"  note: {name} has office-hours rows but is no longer on the grid - rows "
              f"left in place, delete them by hand if they really are gone")

    if dry_run:
        print("dry run - nothing written")
        return
    if not added and not removed and not new_people:
        return

    # -- people: a name the workbook has never seen ------------------------
    for name in new_people:
        row = people.max_row + 1
        people.cell(row, p_name, name)
        role_cell = people.cell(row, p_role, CONFIRMED_ROLES.get(name) or GUESSED_ROLES.get(name, "gtf"))
        if name not in CONFIRMED_ROLES:
            # Flagged in the sheet, not in `notes` - that column is shown to students.
            role_cell.fill = PatternFill("solid", fgColor="FFF2CC")
            role_cell.comment = Comment(ROLE_GUESS_COMMENT, "faculty grid")
        known[name] = row

    # -- shifts: rebuild only the rows belonging to grid people ------------
    # Everyone else's office-hours rows are carried across whole, every column of them,
    # so an LA's Teams link or a hand-written note survives untouched.
    width = shifts.max_column
    keep = [[shifts.cell(row, c).value for c in range(1, width + 1)]
            for row in range(2, shifts.max_row + 1)
            if is_office_hours(row) and s(shifts.cell(row, col["name"]).value) not in grid_names]

    block = keep + [[name, day, start, end] + [None] * (width - 4)
                    for name, day, start, end in wanted]
    # Sorted on real clock minutes. The migration sorted the times as text, which filed
    # 9 AM after 10 AM and left a stray early row at the foot of three days; rewriting the
    # block is the moment to put that right, and it keeps re-runs slotting in cleanly.
    block.sort(key=lambda r: key(s(r[1]), s(r[2]), s(r[0])))

    first_tutoring = min((row for row in range(2, shifts.max_row + 1)
                          if not is_office_hours(row)), default=shifts.max_row + 1)
    shifts.delete_rows(2, first_tutoring - 2)
    shifts.insert_rows(2, len(block))
    for i, values in enumerate(block):
        for c, value in enumerate(values, start=1):
            cell = shifts.cell(2 + i, c, value)
            if c in (col["start"], col["end"]):
                cell.alignment = Alignment(horizontal="left")
        # mode, location, courses, active, program stay blank on a new row: in person, in
        # the room settings gives their role, the default course list, active, office hours.

    wb.save(BOOK)
    total = sum(1 for r in range(2, shifts.max_row + 1) if is_office_hours(r))
    print(f"written: {len(wanted)} faculty/GTF shifts, {total} office-hours rows in all")
    print(f"check the site at ?check=1 before pushing")


if __name__ == "__main__":
    main()
